from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import monthly_report


BASE_DIR = Path(__file__).resolve().parent
TARGET_ROWS = [
    "A股",
    "美股",
    "港股",
    "德国",
    "日本",
    "印度",
    "欧洲",
    "法国",
    "黄金",
    "有色金属",
    "豆粕",
    "能化",
    "A债",
    "货币",
]
VALUE_COLUMNS = ["期初占比", "期末占比", "日均占比", "收益贡献"]
OUTPUT_COLUMNS = ["资产收益率", "平均仓位", "收益贡献", "基准指数选取", "指数表现", "超额收益率"]

DEFAULT_BENCHMARKS = {
    "A股": ("中证全指", "000985.CSI"),
    "美股": ("标普500", "SPX.GI"),
    "港股": ("恒生指数（经汇率调整）", "159920.SZ"),
    "德国": ("德国DAX指数", "GDAXI.GI"),
    "日本": ("日经225", "N225.GI"),
    "印度": ("印度SENSEX30", "935600.MI"),
    "欧洲": ("欧洲STOXX600", "SXXP.GI"),
    "法国": ("法国CAC40", "FCHI.GI"),
    "黄金": ("上海金现货价格收益率", "518850.SH"),
    "有色金属": ("上期有色金属指数", "IMCI.SHF"),
    "豆粕": ("大商所豆粕期货价格指数", "DCESMFI.DCE"),
    "能化": ("易盛能化", "000201.CZC"),
    "A债": ("中债-综合全价(总值)指数", "CBA00203.CS"),
    "货币": ("万得货币市场基金指数", "885009.WI"),
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="调用原月报数据库数据生成带基准表现和超额收益的大类资产业绩归因表。"
    )
    parser.add_argument("--date", default="2026-06-17", help="报告日期目录，例如 2026-06-17。")
    parser.add_argument("--fund", default="华夏盈泰稳健", help="基金目录名称。")
    parser.add_argument("--start", default="2026-06-01", help="归因起始日期。")
    parser.add_argument("--end", default="2026-06-17", help="归因结束日期。")
    parser.add_argument(
        "--data-root",
        default=str(BASE_DIR),
        help="数据根目录；脚本读取 <data-root>/<date>/<fund>/ 下的 Excel。",
    )
    parser.add_argument("--output", default=None, help="输出 Excel 文件路径；默认输出到基金目录下。")
    return parser.parse_args()


def find_source_excel(report_dir: Path, start: str, end: str) -> Path:
    files = [path.name for path in report_dir.iterdir() if path.is_file() and not path.name.startswith("~$")]
    excels = monthly_report.extract_excel_dates(files)
    match = excels[(excels["start"] == pd.to_datetime(start)) & (excels["end"] == pd.to_datetime(end))]
    match = match[match.index.to_series().str.contains("持仓收益分析", na=False)]
    if len(match) != 1:
        raise FileNotFoundError(f"在 {report_dir} 中没有唯一匹配 {start}_{end} 的持仓收益分析 Excel。")
    return report_dir / match.index[0]


def load_source_rows(excel_path: Path) -> pd.DataFrame:
    df = pd.read_excel(excel_path, header=3).drop(columns="Unnamed: 0", errors="ignore")
    df = df.iloc[:-2].copy()
    df[["一级分类", "二级分类"]] = df[["一级分类", "二级分类"]].ffill()
    df = df.dropna(subset=["证券代码"]).copy()
    for column in VALUE_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0.0)
    return df


def add_known_suffix(code: str) -> str:
    code = str(code)
    if "." in code:
        return code
    for universe in [monthly_report.returns, monthly_report.index_return, monthly_report.result]:
        if universe is None:
            continue
        index = universe.columns if hasattr(universe, "columns") else universe.index
        full_code = monthly_report.add_tail(code, pd.DataFrame(columns=index))
        if full_code != code:
            return full_code
    for suffix in ["OF", "SH", "SZ", "HK", "BJ"]:
        full_code = f"{code}.{suffix}"
        if full_code in monthly_report.result.index:
            return full_code
    return code


def normalize_category(category: str | float | None, name: str = "") -> str | None:
    if pd.isna(category):
        category = None
    text = f"{category or ''} {name}"
    if any(keyword in text for keyword in ["德国", "DAX"]):
        return "德国"
    if "日本" in text:
        return "日本"
    if "印度" in text:
        return "印度"
    if "欧洲" in text:
        return "欧洲"
    if "法国" in text:
        return "法国"
    if any(keyword in text for keyword in ["美股", "美国", "纳斯达克", "标普", "道琼斯", "环球股票"]):
        return "美股"
    if any(keyword in text for keyword in ["港股", "香港", "恒生", "H股"]):
        return "港股"
    if category in ["A股", "主动权益", "被动权益"]:
        return "A股"
    if category in ["黄金", "黄金ETF"]:
        return "黄金"
    if category in ["有色金属"]:
        return "有色金属"
    if category in ["豆粕"]:
        return "豆粕"
    if category in ["能化"]:
        return "能化"
    if category in ["货币", "货币基金", "现金"]:
        return "货币"
    if any(keyword in text for keyword in ["债券", "短债", "纯债", "固收", "固定收益", "美元收益", "亚洲策略"]):
        return "A债"
    if category in ["A债", "海外债券", "境外固收", "债券", "债券ETF", "纯债债基", "二级债基", "偏债混合"]:
        return "A债"
    return None


def classify_with_db(row: pd.Series, fund_mapper: pd.DataFrame) -> str | None:
    first = str(row["一级分类"])
    second = str(row["二级分类"])
    code = str(row["证券代码"])
    name = str(row["证券名称"])

    if first == "股票":
        return "港股" if "HK" in code else "A股"
    if first == "债券":
        return "A债"
    if first == "现金":
        return "货币" if second in ["活期存款", "备付金保证金"] else None
    if first != "基金":
        return None

    full_code = add_known_suffix(code)
    benchmark_code = monthly_report.result.get(full_code)
    db_category = monthly_report.index_category.get(benchmark_code) if benchmark_code is not None else None
    normalized = normalize_category(db_category, name)
    if normalized is not None:
        return normalized

    mapper_category = None
    if code in fund_mapper.index:
        mapper_category = fund_mapper.loc[code, "资产类型"]
    return normalize_category(mapper_category, name)


def build_classified_detail(df: pd.DataFrame, fund_mapper: pd.DataFrame) -> pd.DataFrame:
    detail = df.copy()
    detail["完整代码"] = detail["证券代码"].astype(str).map(add_known_suffix)
    detail["基金基准代码"] = detail["完整代码"].map(monthly_report.result)
    detail["数据库资产分类"] = detail["基金基准代码"].map(monthly_report.index_category)
    detail["归因资产类型"] = detail.apply(classify_with_db, axis=1, fund_mapper=fund_mapper)
    return detail


def benchmark_return(code: str, start: str, end: str) -> float:
    returns = pd.concat([monthly_report.index_return, monthly_report.returns], axis=1)
    returns = returns.loc[:, ~returns.columns.duplicated()]
    if code not in returns.columns:
        return np.nan
    series = returns.loc[pd.to_datetime(start) : pd.to_datetime(end), code].dropna()
    if series.empty:
        return np.nan
    return (series + 1).prod() - 1


def add_benchmark_columns(table: pd.DataFrame, start: str, end: str) -> pd.DataFrame:
    result = table.copy()
    for asset, (name, code) in DEFAULT_BENCHMARKS.items():
        result.loc[asset, "基准指数选取"] = name
        result.loc[asset, "指数表现"] = benchmark_return(code, start, end)
    result["超额收益率"] = result["资产收益率"] - result["指数表现"]
    return result[OUTPUT_COLUMNS]


def build_target_table(detail: pd.DataFrame, start: str, end: str) -> pd.DataFrame:
    grouped = (
        detail.dropna(subset=["归因资产类型"])
        .groupby("归因资产类型")[VALUE_COLUMNS]
        .sum()
        .reindex(TARGET_ROWS, fill_value=0.0)
    )
    table = pd.DataFrame(index=TARGET_ROWS)
    table.index.name = ""
    table["资产收益率"] = grouped["收益贡献"] / grouped["日均占比"]
    table.loc[grouped["日均占比"].abs() <= 0.00001, "资产收益率"] = np.nan
    table["平均仓位"] = grouped["日均占比"]
    table["收益贡献"] = grouped["收益贡献"]
    return add_benchmark_columns(table, start, end)


def build_original_project_table(excel_path: Path, fund_mapper: pd.DataFrame, start: str, end: str) -> pd.DataFrame:
    month_table = monthly_report.analyse_table(str(excel_path), fund_mapper, monthly_report.result)
    table = pd.DataFrame(index=month_table.index)
    table["资产收益率"] = month_table["收益贡献"] / month_table["平均仓位"]
    table.loc[month_table["平均仓位"].abs() <= 0.00001, "资产收益率"] = np.nan
    table["平均仓位"] = month_table["平均仓位"]
    table["收益贡献"] = month_table["收益贡献"]

    bench = monthly_report.bench.set_index("资产类型")
    for asset, row in bench.iterrows():
        if asset not in table.index:
            continue
        table.loc[asset, "基准指数选取"] = row["基准指数名称"]
        table.loc[asset, "指数表现"] = benchmark_return(row["基准指数代码"], start, end)
    table["超额收益率"] = table["资产收益率"] - table["指数表现"]
    return table[OUTPUT_COLUMNS]


def write_output(output_path: Path, sheets: dict[str, pd.DataFrame]):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=sheet_name != "分类明细")

        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        for ws in writer.book.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="宋体", size=11, bold=cell.row == 1 or cell.column == 1)
                    if cell.row == 1:
                        cell.fill = PatternFill("solid", fgColor="FFFFFF")
            for column in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(column)].width = 16
            if ws.title != "分类明细":
                for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.00%"


def main():
    args = parse_args()
    monthly_report.init_report_data()

    report_dir = Path(args.data_root) / args.date / args.fund
    excel_path = find_source_excel(report_dir, args.start, args.end)
    fund_mapper = monthly_report.build_fund_mapper(
        monthly_report.result,
        monthly_report.index_category,
        monthly_report.fund_basic,
        monthly_report.benches,
    )

    source_rows = load_source_rows(excel_path)
    detail = build_classified_detail(source_rows, fund_mapper)
    target_table = build_target_table(detail, args.start, args.end)
    original_table = build_original_project_table(excel_path, fund_mapper, args.start, args.end)

    output_path = (
        Path(args.output)
        if args.output
        else report_dir / f"{args.fund}_{args.start}_{args.end}_资产业绩归因表_DB版.xlsx"
    )
    write_output(
        output_path,
        {
            "资产业绩归因_DB地区口径": target_table,
            "原项目月报口径": original_table,
            "分类明细": detail,
        },
    )
    print(f"已生成：{output_path}")
    print(target_table.applymap(lambda x: "" if pd.isna(x) else f"{x:.2%}" if isinstance(x, float) else x).to_string())


if __name__ == "__main__":
    main()
