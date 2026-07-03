from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_REPORT_DIR = BASE_DIR / "data" / "2026-06-26" / "华夏聚诚优选"
DEFAULT_START = "2026-04-17"
DEFAULT_END = "2026-06-26"

VALUE_COLUMNS = ["期初占比", "期末占比", "日均占比", "收益贡献"]
OUTPUT_COLUMNS = ["资产收益率", "平均仓位", "收益贡献"]
TARGET_ROWS = [
    "A股",
    "美股",
    "港股",
    "德国",
    "日本",
    "印度",
    "欧洲",
    "法国",
    "英国",
    "越南",
    "黄金",
    "有色金属",
    "豆粕",
    "能化",
    "A债",
    "货币",
]

# 本地文件没有数据库里的基金基准映射，少数跨境/商品基金用代码兜底。
CODE_CATEGORY_OVERRIDES = {
    "001061": "A债",
    "002400": "A债",
    "968132": "A债",
    "968115": "A债",
    "968114": "A债",
    "968157": "美股",
    "518850": "黄金",
    "518880": "黄金",
    "159980": "有色金属",
    "159981": "能化",
    "159985": "豆粕",
    "513300": "美股",
    "159655": "美股",
    "513400": "美股",
    "513030": "德国",
    "006105": "印度",
    "007280": "日本",
    "006282": "欧洲",
    "159920": "港股",
    "513330": "港股",
    "513180": "港股",
    "159726": "港股",
    "513910": "港股",
}


def parse_args():
    parser = argparse.ArgumentParser(description="仅使用本地持仓收益分析 Excel 生成资产归因表。")
    parser.add_argument("--report-dir", default=str(DEFAULT_REPORT_DIR), help="持仓收益分析 Excel 所在目录。")
    parser.add_argument("--start", default=DEFAULT_START, help="区间开始日期，例如 2025-10-24。")
    parser.add_argument("--end", default=DEFAULT_END, help="区间结束日期，例如 2026-06-26。")
    parser.add_argument("--output", default=None, help="输出 Excel 路径；默认输出到 report-dir 下。")
    return parser.parse_args()


def extract_excel_dates(files: list[str]) -> pd.DataFrame:
    pattern = r"(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})"
    records = {}
    for file_name in files:
        if not file_name.endswith(".xlsx") or file_name.startswith("~$"):
            continue
        match = re.search(pattern, file_name)
        if match:
            records[file_name] = {"start": match.group(1), "end": match.group(2)}
    result = pd.DataFrame.from_dict(records, orient="index")
    if result.empty:
        return result
    result["start"] = pd.to_datetime(result["start"])
    result["end"] = pd.to_datetime(result["end"])
    return result.sort_values(["start", "end"])


def find_source_excel(report_dir: Path, start: str, end: str) -> Path:
    excels = extract_excel_dates([path.name for path in report_dir.iterdir() if path.is_file()])
    if excels.empty:
        raise FileNotFoundError(f"{report_dir} 下没有可识别日期区间的 xlsx 文件。")
    match = excels[
        (excels["start"] == pd.to_datetime(start))
        & (excels["end"] == pd.to_datetime(end))
        & excels.index.to_series().str.contains("持仓收益分析", na=False)
    ]
    if len(match) != 1:
        raise FileNotFoundError(f"在 {report_dir} 中没有唯一匹配 {start}_{end} 的持仓收益分析 Excel。")
    return report_dir / match.index[0]


def load_source_rows(excel_path: Path) -> pd.DataFrame:
    df = pd.read_excel(excel_path, header=3).drop(columns="Unnamed: 0", errors="ignore")
    df = df.iloc[:-2].copy()
    df[["一级分类", "二级分类"]] = df[["一级分类", "二级分类"]].ffill()
    df = df.dropna(subset=["证券代码"]).copy()
    for column in VALUE_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0.0)
    df["证券代码"] = df["证券代码"].astype(str)
    df["证券名称"] = df["证券名称"].fillna("").astype(str)
    return df


def classify_row(row: pd.Series) -> str | None:
    first = str(row["一级分类"])
    second = str(row["二级分类"])
    code = str(row["证券代码"]).split(".")[0]
    name = str(row["证券名称"])
    text = f"{second} {name}"

    if code in CODE_CATEGORY_OVERRIDES:
        return CODE_CATEGORY_OVERRIDES[code]
    if first == "股票":
        return "港股" if "HK" in code.upper() else "A股"
    if first == "债券":
        return "A债"
    if first == "回购":
        return "货币"
    if first == "现金":
        return "货币" if second in ["活期存款", "备付金保证金"] else None
    if first != "基金":
        return None

    if "货币" in second or "货币" in name:
        return "货币"
    if second in ["纯债及一级债基", "二级债基"] or any(
        keyword in text for keyword in ["短债", "纯债", "债券", "固收", "固定收益", "美元收益", "亚洲策略"]
    ):
        return "A债"
    if any(keyword in text for keyword in ["黄金", "上海金"]):
        return "黄金"
    if "豆粕" in text:
        return "豆粕"
    if any(keyword in text for keyword in ["能源化工期货", "易盛郑商所能源化工", "能化"]):
        return "能化"
    if "有色金属" in text or "稀有金属" in text:
        return "有色金属"
    if any(keyword in text for keyword in ["德国", "DAX"]):
        return "德国"
    if "日本" in text:
        return "日本"
    if "印度" in text:
        return "印度"
    if any(keyword in text for keyword in ["法国", "CAC"]):
        return "法国"
    if any(keyword in text for keyword in ["英国", "富时", "FTSE"]):
        return "英国"
    if "越南" in text:
        return "越南"
    if any(keyword in text for keyword in ["欧洲", "STOXX"]):
        return "欧洲"
    if any(keyword in text for keyword in ["港股", "香港", "恒生", "H股"]):
        return "港股"
    if any(keyword in text for keyword in ["美股", "美国", "纳斯达克", "标普", "道琼斯", "环球股票"]):
        return "美股"
    if second == "偏股基金" or "ETF" in name or "股票" in name:
        return "A股"
    return None


def build_detail(df: pd.DataFrame) -> pd.DataFrame:
    detail = df.copy()
    detail["归因资产类型"] = detail.apply(classify_row, axis=1)
    return detail


def build_attribution_table(detail: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        detail.dropna(subset=["归因资产类型"])
        .groupby("归因资产类型")[VALUE_COLUMNS]
        .sum()
        .reindex(TARGET_ROWS, fill_value=0.0)
    )
    table = pd.DataFrame(index=TARGET_ROWS)
    table.index.name = ""
    table["资产收益率"] = grouped["收益贡献"] / grouped["日均占比"]
    table.loc[grouped["日均占比"].abs() <= 0.00001, "资产收益率"] = np.nan
    table["平均仓位"] = grouped["日均占比"]
    table["收益贡献"] = grouped["收益贡献"]
    table = table.loc[(table["平均仓位"].abs() > 0.00001) | (table["收益贡献"].abs() > 0.000001)]
    return table[OUTPUT_COLUMNS]


def write_output(output_path: Path, table: pd.DataFrame, detail: pd.DataFrame):
    detail_columns = ["归因资产类型", "一级分类", "二级分类", "证券代码", "证券名称"] + VALUE_COLUMNS
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="资产归因表")
        detail[detail_columns].to_excel(writer, sheet_name="分类明细", index=False)

        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        for ws in writer.book.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="宋体", size=11, bold=cell.row == 1 or cell.column == 1)
                    if cell.row == 1:
                        cell.fill = PatternFill("solid", fgColor="E7E6E6")
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 16
            if ws.title == "资产归因表":
                for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.00%"
                ws.column_dimensions["A"].width = 12
            else:
                ws.column_dimensions["E"].width = 36
                for row in ws.iter_rows(min_row=2, min_col=6, max_col=ws.max_column):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.00%"


def main():
    args = parse_args()
    report_dir = Path(args.report_dir)
    source = find_source_excel(report_dir, args.start, args.end)
    output = Path(args.output) if args.output else report_dir / f"本地资产归因表_{args.start}_{args.end}.xlsx"

    rows = load_source_rows(source)
    detail = build_detail(rows)
    table = build_attribution_table(detail)
    write_output(output, table, detail)

    print(f"源文件：{source}")
    print(f"输出：{output}")
    print(table.map(lambda x: f"{x:.2%}" if pd.notna(x) else "").to_string())


if __name__ == "__main__":
    main()
